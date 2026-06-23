from .utils import login_ofertas_required
from ofertas.sicredi_service import buscar_pix_sicredi
from datetime import datetime, timedelta, timezone
from sqlalchemy.orm import sessionmaker
from models import OfertaRecebida

from flask import (
    Blueprint,
    render_template,
    request,
    redirect,
    url_for,
    flash,
    send_file
)

from extensions import db

from models import (
    Comunidade,
    ComunidadePix,
    TipoArrecadacao,
    OfertaRecebida
)

ofertas_bp = Blueprint(
    "ofertas",
    __name__,
    url_prefix="/ofertas"
)


@ofertas_bp.route("/comunidades")
@login_ofertas_required
def comunidades():

    if session.get("administrador"):

        comunidades = Comunidade.query.order_by(
            Comunidade.nome
        ).all()

    else:

        comunidades = Comunidade.query.filter_by(
            id=session["comunidade_id"]
        ).all()

    return render_template(

         "ofertas/comunidade_pix.html",

        comunidades=comunidades

    )

from models import Comunidade, TipoArrecadacao

@ofertas_bp.route(
    "/comunidades/novo",
    methods=["GET","POST"]
)
@login_ofertas_required
def comunidade_nova():

    tipos = TipoArrecadacao.query.filter_by(
        ativo=True
    ).order_by(
        TipoArrecadacao.descricao
    ).all()

    if request.method == "POST":

        c = Comunidade()

        c.nome = request.form["nome"]
        c.codigo = request.form["codigo"]
        c.responsavel = request.form["responsavel"]
        c.telefone = request.form["telefone"]
        c.cor = request.form["cor"]
        c.ativa = request.form.get("ativa") == "1"
        c.senha_acesso = request.form["senha_acesso"]
        c.administrador = (request.form.get("administrador") == "1")
        c.acesso_restrito = (request.form.get("acesso_restrito") == "1")        

        db.session.add(c)
        db.session.commit()

        return redirect(url_for("ofertas.comunidades"))

    return render_template(
        "ofertas/comunidade_form.html",
        tipos=tipos
    )



@ofertas_bp.route(
    "/comunidades/<int:id>",
    methods=["GET", "POST"]
)
@login_ofertas_required
def editar_comunidade(id):

    comunidade = Comunidade.query.get_or_404(id)

    if (
        not session.get("administrador")
        and
        comunidade.id != session["comunidade_id"]
    ):

        flash(
            "Acesso negado.",
            "danger"
        )

        return redirect(
            url_for("ofertas.inicio")
        )

    if request.method == "POST":

        comunidade.nome = request.form["nome"]
        comunidade.codigo = request.form["codigo"]
        comunidade.responsavel = request.form["responsavel"]
        comunidade.telefone = request.form["telefone"]
        comunidade.cor = request.form["cor"]
        comunidade.ativa = (request.form.get("ativa") == "1")
        comunidade.senha_acesso = request.form["senha_acesso"]
        comunidade.administrador = (request.form.get("administrador") == "1")
        comunidade.acesso_restrito = (request.form.get("acesso_restrito") == "1")        

        db.session.commit()

        flash(
            "Comunidade alterada com sucesso.",
            "success"
        )

        return redirect(
            url_for(
                "ofertas.comunidades"
            )
        )

    tipos = TipoArrecadacao.query.filter_by(
        ativo=True
    ).order_by(
        TipoArrecadacao.descricao
    ).all()

    pix = ComunidadePix.query.filter_by(
        comunidade_id=comunidade.id
    ).order_by(
        ComunidadePix.descricao
    ).all()

    return render_template(

        "ofertas/comunidade_form.html",

        comunidade=comunidade,

        tipos=tipos,

        pix=pix

    )

@ofertas_bp.route(
    "/comunidades/excluir/<int:id>"
)
@login_ofertas_required
def excluir_comunidade(id):

    c = Comunidade.query.get_or_404(id)

    if (
        not session.get("administrador")
        and
        c.id != session["comunidade_id"]
    ):

        flash(
            "Acesso negado.",
            "danger"
        )

        return redirect(
            url_for("ofertas.inicio")
        )

    db.session.delete(c)

    db.session.commit()

    flash("Comunidade excluída.")

    return redirect(
        url_for("ofertas.comunidades")
    )


import qrcode


#----------------------------------------
# CRC16
#----------------------------------------

def crc16(payload):

    crc = 0xFFFF

    for c in payload:

        crc ^= ord(c) << 8

        for _ in range(8):

            if crc & 0x8000:

                crc = (
                    (crc << 1)
                    ^
                    0x1021
                ) & 0xFFFF

            else:

                crc = (
                    crc << 1
                ) & 0xFFFF

    return format(crc, "04X")


#----------------------------------------
# TAG
#----------------------------------------

def tag(id, valor):

    return (
        id +
        str(len(valor)).zfill(2) +
        valor
    )


#----------------------------------------
# GERA PIX
#----------------------------------------

def gerar_pix(
        chave,
        nome,
        cidade,
        identificacao):

    merchant = ""

    merchant += tag(
        "00",
        "BR.GOV.BCB.PIX"
    )

    merchant += tag(
        "01",
        chave
    )

    payload = ""

    payload += tag(
        "00",
        "01"
    )

    payload += tag(
        "01",
        "11"
    )

    payload += tag(
        "26",
        merchant
    )

    payload += tag(
        "52",
        "0000"
    )

    payload += tag(
        "53",
        "986"
    )

    payload += tag(
        "58",
        "BR"
    )

    payload += tag(
        "59",
        nome[:25]
    )

    payload += tag(
        "60",
        cidade[:15]
    )

    adicional = tag(
        "05",
        identificacao
    )

    payload += tag(
        "62",
        adicional
    )

    payload += "6304"

    payload += crc16(payload)

    return payload

# -----------------------
# QRCode
# -----------------------

@ofertas_bp.route("/qrcode/<int:id>")
@login_ofertas_required
def gerar_qrcode(id):

    import os

    comunidade = ComunidadePix.query.get_or_404(id)

    payload = gerar_pix(

        chave=comunidade.chave_pix,

        nome="PAROQUIA NS APARECIDA",

        cidade="PALMAS",

        identificacao=comunidade.txid

    )

    qr = qrcode.QRCode(

        version=None,

        error_correction=qrcode.constants.ERROR_CORRECT_M,

        box_size=10,

        border=4

    )

    qr.add_data(payload)

    qr.make(fit=True)

    img = qr.make_image(

        fill_color="black",

        back_color="white"

    )

    os.makedirs(

        "static/qrcodes",

        exist_ok=True

    )

    caminho = os.path.join(

        "static",

        "qrcodes",

        f"{comunidade.txid}.png"

    )

    img.save(caminho)

    return send_file(

        caminho,

        mimetype="image/png",

        as_attachment=False

    )


@ofertas_bp.route("/recebidas")
@login_ofertas_required
def ofertas():

    consulta = OfertaRecebida.query

    if not session.get("administrador"):

        consulta = consulta.filter(
            OfertaRecebida.comunidade_id ==
            session["comunidade_id"]
        )

    lista = consulta.order_by(
        OfertaRecebida.datahora.desc()
    ).all()    

    return render_template(

        "ofertas/ofertas.html",

        lista=lista

    )

@ofertas_bp.route("/relatorios", methods=["GET", "POST"])
@login_ofertas_required
def relatorios():

    comunidades = Comunidade.query.order_by(
        Comunidade.nome
    ).all()

    tipos = TipoArrecadacao.query.order_by(
        TipoArrecadacao.descricao
    ).all()

    consulta = OfertaRecebida.query

    if not session.get(
        "administrador"
    ):

        consulta = consulta.filter(
            OfertaRecebida.comunidade_id==
            session["comunidade_id"]

        )

    if request.method == "POST":

        data_inicial = request.form.get("data_inicial")
        data_final = request.form.get("data_final")
        comunidade = request.form.get("comunidade")
        tipo = request.form.get("tipo")

        if data_inicial:
            consulta = consulta.filter(
                OfertaRecebida.datahora >= data_inicial
            )

        if data_final:
            consulta = consulta.filter(
                OfertaRecebida.datahora <=
                data_final + " 23:59:59"
            )

        if comunidade:
            consulta = consulta.filter(
                OfertaRecebida.comunidade_id == comunidade
            )

        if tipo:
            consulta = consulta.filter(
                OfertaRecebida.tipo_id == tipo
            )

    lista = consulta.order_by(
        OfertaRecebida.datahora.desc()
    ).all()

    total = sum(
        float(x.valor)
        for x in lista
    )

    return render_template(

        "ofertas/relatorios.html",

        lista=lista,

        total=total,

        comunidades=comunidades,

        tipos=tipos

    )

@ofertas_bp.route("/imprimir_qrcode/<int:id>")
@login_ofertas_required
def imprimir_qrcode(id):

    comunidade = ComunidadePix.query.get_or_404(id)

    return render_template(

        "ofertas/imprimir_qrcode.html",

        comunidade=comunidade

    )

from ofertas.services import importar_pix_automatico as importar_pix_service

@ofertas_bp.route("/importar_pix")
@login_ofertas_required
def importar_pix():

    total = importar_pix_service()

    flash(
        f"✅ {total} PIX importados com sucesso!",
        "success"
    )

    return redirect(
        url_for("ofertas.ofertas")
    )

@ofertas_bp.route("/relatorio_comunidade", methods=["GET", "POST"])
@login_ofertas_required
def relatorio_comunidade():

    consulta = db.session.query(

        Comunidade.nome,

        db.func.count(OfertaRecebida.id).label("quantidade"),

        db.func.sum(OfertaRecebida.valor).label("total")

    ).join(

        OfertaRecebida,

        OfertaRecebida.comunidade_id == Comunidade.id

    )

    if not session.get("administrador"):

        consulta = consulta.filter(
            OfertaRecebida.comunidade_id ==
            session["comunidade_id"]
        )    

    if request.method == "POST":

        data_inicial = request.form.get("data_inicial")
        data_final = request.form.get("data_final")

        if data_inicial:

            consulta = consulta.filter(
                OfertaRecebida.datahora >= data_inicial
            )

        if data_final:

            consulta = consulta.filter(
                OfertaRecebida.datahora <= data_final + " 23:59:59"
            )

    lista = consulta.group_by(
        Comunidade.nome
    ).order_by(
        Comunidade.nome
    ).all()

    total_geral = sum(
        float(x.total or 0)
        for x in lista
    )

    return render_template(

        "ofertas/relatorio_comunidade.html",

        lista=lista,

        total_geral=total_geral

    )


from sqlalchemy import func

@ofertas_bp.route("/dashboard", methods=["GET", "POST"])
@login_ofertas_required
def dashboard():

    data_inicial = None
    data_final = None
    comunidade_id = None

    consulta = OfertaRecebida.query

    if not session.get(
        "administrador"
    ):

        consulta = consulta.filter(

            OfertaRecebida.comunidade_id==

            session["comunidade_id"]

        )

    comunidades_lista = Comunidade.query.order_by(
        Comunidade.nome
    ).all()

    if request.method == "POST":

        data_inicial = request.form.get("data_inicial")
        data_final = request.form.get("data_final")
        comunidade_id = request.form.get("comunidade")

        if data_inicial:

            consulta = consulta.filter(
                OfertaRecebida.datahora >=
                datetime.strptime(
                    data_inicial,
                    "%Y-%m-%d"
                )
            )

        if data_final:

            consulta = consulta.filter(
                OfertaRecebida.datahora <=
                datetime.strptime(
                    data_final + " 23:59:59",
                    "%Y-%m-%d %H:%M:%S"
                )
            )

        if comunidade_id:

            consulta = consulta.filter(
                OfertaRecebida.comunidade_id == int(comunidade_id)
            )

    total = sum(
        float(x.valor)
        for x in consulta.all()
    )

    quantidade = consulta.count()

    ranking = db.session.query(

        Comunidade.nome,

        db.func.sum(
            OfertaRecebida.valor
        ).label("total")

    ).join(

        OfertaRecebida,

        OfertaRecebida.comunidade_id == Comunidade.id

    )

    if not session.get("administrador"):

        ranking = ranking.filter(

            OfertaRecebida.comunidade_id ==

            session["comunidade_id"]

        )    

    if data_inicial:

        ranking = ranking.filter(
            OfertaRecebida.datahora >=
            datetime.strptime(
                data_inicial,
                "%Y-%m-%d"
            )
        )

    if data_final:

        ranking = ranking.filter(
            OfertaRecebida.datahora <=
            datetime.strptime(
                data_final + " 23:59:59",
                "%Y-%m-%d %H:%M:%S"
            )
        )

    if comunidade_id:

        ranking = ranking.filter(
            OfertaRecebida.comunidade_id == int(comunidade_id)
        )

    ranking = ranking.group_by(
        Comunidade.nome
    ).order_by(
        db.func.sum(
            OfertaRecebida.valor
        ).desc()
    ).all()

    return render_template(

        "ofertas/dashboard.html",

        total=total,

        quantidade=quantidade,

        comunidades=ranking,

        lista_comunidades=comunidades_lista,

        data_inicial=data_inicial,

        data_final=data_final,

        comunidade_id=comunidade_id

    )



@ofertas_bp.route("/ranking", methods=["GET", "POST"])
@login_ofertas_required
def ranking_comunidades():

    data_inicial = None
    data_final = None

    consulta = db.session.query(


        Comunidade.id,

        Comunidade.nome,

        db.func.count(
            OfertaRecebida.id
        ).label("qtd_pix"),

        db.func.sum(
            OfertaRecebida.valor
        ).label("total")

    ).join(

        OfertaRecebida,

        OfertaRecebida.comunidade_id == Comunidade.id

    )

    if not session.get("administrador"):

        consulta = consulta.filter(
            OfertaRecebida.comunidade_id ==
            session["comunidade_id"]
        )    

    if request.method == "POST":

        data_inicial = request.form.get("data_inicial")
        data_final = request.form.get("data_final")

        if data_inicial:

            consulta = consulta.filter(

                OfertaRecebida.datahora >=

                datetime.strptime(
                    data_inicial,
                    "%Y-%m-%d"
                )

            )

        if data_final:

            consulta = consulta.filter(

                OfertaRecebida.datahora <=

                datetime.strptime(
                    data_final + " 23:59:59",
                    "%Y-%m-%d %H:%M:%S"
                )

            )

    ranking = consulta.group_by(

        Comunidade.id,
        Comunidade.nome

    ).order_by(

        db.func.sum(
            OfertaRecebida.valor
        ).desc()

    ).all()

    total_geral = sum(
        float(x.total or 0)
        for x in ranking
    )

    return render_template(

        "ofertas/ranking.html",

        ranking=ranking,

        total_geral=total_geral,

        data_inicial=data_inicial,

        data_final=data_final

    )

# ============================
# EXPORTAR EXCEL
# ============================

from flask import request, send_file
from openpyxl import Workbook
from io import BytesIO


@ofertas_bp.route("/exportar_excel")
@login_ofertas_required
def exportar_excel():

    data_inicial = request.args.get("data_inicial")
    data_final = request.args.get("data_final")
    comunidade_id = request.args.get("comunidade")

    consulta = OfertaRecebida.query



    if not session.get(
        "administrador"
    ):

        consulta = consulta.filter(

            OfertaRecebida.comunidade_id==

            session["comunidade_id"]

        )

    if data_inicial:

        consulta = consulta.filter(

            OfertaRecebida.datahora >=

            datetime.strptime(
                data_inicial,
                "%Y-%m-%d"
            )

        )

    if data_final:

        consulta = consulta.filter(

            OfertaRecebida.datahora <=

            datetime.strptime(
                data_final + " 23:59:59",
                "%Y-%m-%d %H:%M:%S"
            )

        )

    if comunidade_id:

        consulta = consulta.filter(

            OfertaRecebida.comunidade_id == int(comunidade_id)

        )

    lista = consulta.order_by(
        OfertaRecebida.datahora
    ).all()

    wb = Workbook()

    ws = wb.active

    ws.title = "Ofertas"

    ws.append([
        "Data/Hora",
        "Pagador",
        "EndToEndId",
        "Valor"
    ])

    total = 0

    for o in lista:

        total += float(o.valor)

        ws.append([

            o.datahora.strftime(
                "%d/%m/%Y %H:%M"
            ),

            o.comunidade.nome if o.comunidade else "",

            o.tipo.descricao if o.tipo else "",

            o.txid,

            float(o.valor)

        ])

    ws.append([])

    ws.append([
        "",
        "",
        "",
        "TOTAL",
        total
    ])

    arquivo = BytesIO()

    wb.save(arquivo)

    arquivo.seek(0)

    return send_file(

        arquivo,

        as_attachment=True,

        download_name="Prestacao_Contas.xlsx",

        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    )


# ============================
# PDF
# ============================

from reportlab.platypus import (
    SimpleDocTemplate,
    Table,
    TableStyle
)

from reportlab.lib import colors


@ofertas_bp.route("/imprimir_pdf")
@login_ofertas_required
def imprimir_pdf():

    data_inicial = request.args.get("data_inicial")
    data_final = request.args.get("data_final")
    comunidade_id = request.args.get("comunidade")

    consulta = OfertaRecebida.query

    if not session.get(
        "administrador"
    ):

        consulta = consulta.filter(

            OfertaRecebida.comunidade_id==

            session["comunidade_id"]

        )

    if data_inicial:

        consulta = consulta.filter(

            OfertaRecebida.datahora >=

            datetime.strptime(
                data_inicial,
                "%Y-%m-%d"
            )

        )

    if data_final:

        consulta = consulta.filter(

            OfertaRecebida.datahora <=

            datetime.strptime(
                data_final + " 23:59:59",
                "%Y-%m-%d %H:%M:%S"
            )

        )

    if comunidade_id:

        consulta = consulta.filter(

            OfertaRecebida.comunidade_id == int(comunidade_id)

        )

    lista = consulta.order_by(
        OfertaRecebida.datahora
    ).all()

    buffer = BytesIO()

    doc = SimpleDocTemplate(buffer)

    dados = [[

        "Data",

        "Comunidade",

        "Tipo",

        "ID",

        "Valor"

    ]]

    total = 0

    for o in lista:

        total += float(o.valor)

        dados.append([

            o.datahora.strftime(
                "%d/%m/%Y %H:%M"
            ),

            o.comunidade.nome if o.comunidade else "",

            o.tipo.descricao if o.tipo else "",

            o.endtoendid,

            f"R$ {o.valor:.2f}"

        ])

    dados.append([

        "",

        "",

        "",
        

        "TOTAL",

        f"R$ {total:.2f}"

    ])

    tabela = Table(dados)

    tabela.setStyle(TableStyle([

        ("GRID",(0,0),(-1,-1),1,colors.black),

        ("BACKGROUND",(0,0),(-1,0),colors.lightgrey),

        ("BACKGROUND",(0,-1),(-1,-1),colors.lightgreen),

        ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),

        ("FONTNAME",(0,-1),(-1,-1),"Helvetica-Bold"),

        ("ALIGN",(4,1),(4,-1),"RIGHT"),

        ("FONTSIZE",(0,0),(-1,-1),9)

    ]))

    doc.build([tabela])

    buffer.seek(0)

    return send_file(

        buffer,

        as_attachment=False,

        download_name="Prestacao_Contas.pdf",

        mimetype="application/pdf"

    )

#Rotina IMportação Pix
from ofertas.services import importar_pix_automatico

@ofertas_bp.route("/importar_pix_automatico")
@login_ofertas_required
def importar_pix_automatico():

    total = importar_pix_automatico()

    return f"{total} PIX importados."


@ofertas_bp.route("/", methods=["GET"])
@login_ofertas_required
def inicio():

    return render_template(
        "ofertas/index.html"
    )

@ofertas_bp.route("/comunidades_pix")
@login_ofertas_required
def comunidades_pix():

    comunidades = Comunidade.query.order_by(

        Comunidade.nome

    ).all()

    return render_template(

        "ofertas/comunidades_pix.html",

        comunidades=comunidades

    )

@ofertas_bp.route(
    "/pix/novo",
    methods=["GET","POST"]
)
@login_ofertas_required
def comunidade_pix_novo():

    comunidades = Comunidade.query.order_by(
        Comunidade.nome
    ).all()

    tipos = TipoArrecadacao.query.filter_by(
        ativo=True
    ).order_by(
        TipoArrecadacao.descricao
    ).all()

    comunidade_id = request.args.get(
        "comunidade",
        type=int
    )

    if request.method == "POST":

        c = ComunidadePix()

        c.comunidade_id = int(
            request.form["comunidade_id"]
        )

        c.tipo_id = int(
            request.form["tipo_id"]
        )

        c.descricao = request.form[
            "descricao"
        ]

        c.txid = request.form[
            "txid"
        ].upper()

        c.chave_pix = request.form[
            "chave_pix"
        ]

        c.ativo = (
            request.form.get("ativo") == "1"
        )

        db.session.add(c)

        db.session.commit()

        flash(
            "PIX cadastrado com sucesso.",
            "success"
        )

        return redirect(

            url_for(

                "ofertas.editar_comunidade",

                id=c.comunidade_id

            )

        )

    return render_template(

        "ofertas/comunidade_pix_form.html",

        comunidades=comunidades,

        tipos=tipos,

        comunidade_id=comunidade_id,

        registro=None,

        editando=False

    )

@ofertas_bp.route(
    "/pix/<int:id>",
    methods=["GET", "POST"]
)
@login_ofertas_required
def editar_pix(id):

    registro = ComunidadePix.query.get_or_404(id)

    if (
        not session.get("administrador")
        and
        registro.comunidade_id != session["comunidade_id"]
    ):

        flash(
            "Acesso negado.",
            "danger"
        )

        return redirect(
            url_for("ofertas.inicio")
        )    

    comunidades = Comunidade.query.order_by(
        Comunidade.nome
    ).all()

    tipos = TipoArrecadacao.query.filter_by(
        ativo=True
    ).order_by(
        TipoArrecadacao.descricao
    ).all()

    if request.method == "POST":

        registro.comunidade_id = int(
            request.form["comunidade_id"]
        )

        registro.tipo_id = int(
            request.form["tipo_id"]
        )

        registro.descricao = request.form[
            "descricao"
        ].strip()

        registro.txid = request.form[
            "txid"
        ].strip().upper()

        registro.chave_pix = request.form[
            "chave_pix"
        ].strip()

        registro.ativo = (
            request.form.get("ativo") == "1"
        )

        db.session.commit()

        flash(
            "PIX alterado com sucesso.",
            "success"
        )

        return redirect(

            url_for(

                "ofertas.editar_comunidade",

                id=registro.comunidade_id

            )

        )

    return render_template(

        "ofertas/comunidade_pix_form.html",

        registro=registro,

        comunidades=comunidades,

        tipos=tipos,

        editando=True

    )

@ofertas_bp.route(
    "/pix/excluir/<int:id>"
)
@login_ofertas_required
def excluir_pix(id):

    registro = ComunidadePix.query.get_or_404(id)

    if (
        not session.get("administrador")
        and
        registro.comunidade_id != session["comunidade_id"]
    ):

        flash(
            "Acesso negado.",
            "danger"
        )

        return redirect(
            url_for("ofertas.inicio")
        )


    comunidade_id = registro.comunidade_id

    db.session.delete(registro)

    db.session.commit()

    flash("PIX excluído.")

    return redirect(

        url_for(

            "ofertas.editar_comunidade",

            id=comunidade_id

        )

    )

@ofertas_bp.route(
    "/comunidade/<int:id>/pix"
)
@login_ofertas_required
def comunidade_pix_grid(id):

    pix = ComunidadePix.query.filter_by(

        comunidade_id=id

    ).order_by(

        ComunidadePix.descricao

    ).all()

    return render_template(

        "ofertas/pix_grid.html",

        pix=pix

    )

from sqlalchemy import func

@ofertas_bp.route(
    "/relatorio_ofertas",
    methods=["GET", "POST"]
)
@login_ofertas_required
def relatorio_ofertas():

    comunidades = Comunidade.query.order_by(
        Comunidade.nome
    ).all()

    tipos = TipoArrecadacao.query.order_by(
        TipoArrecadacao.descricao
    ).all()

    consulta = OfertaRecebida.query

    if not session.get(
        "administrador"
    ):

        consulta = consulta.filter(

            OfertaRecebida.comunidade_id==

            session["comunidade_id"]

        )

    data_inicial = None
    data_final = None
    comunidade_id = None
    tipo_id = None

    if request.method == "POST":

        data_inicial = request.form.get(
            "data_inicial"
        )

        data_final = request.form.get(
            "data_final"
        )

        comunidade_id = request.form.get(
            "comunidade"
        )

        tipo_id = request.form.get(
            "tipo"
        )

        if data_inicial:

            consulta = consulta.filter(

                OfertaRecebida.datahora >=

                datetime.strptime(
                    data_inicial,
                    "%Y-%m-%d"
                )

            )

        if data_final:

            consulta = consulta.filter(

                OfertaRecebida.datahora <=

                datetime.strptime(
                    data_final + " 23:59:59",
                    "%Y-%m-%d %H:%M:%S"
                )

            )

        if comunidade_id:

            consulta = consulta.filter(
                OfertaRecebida.comunidade_id ==
                int(comunidade_id)
            )

        if tipo_id:

            consulta = consulta.filter(
                OfertaRecebida.tipo_id ==
                int(tipo_id)
            )

    lista = consulta.order_by(

        OfertaRecebida.comunidade_id,

        OfertaRecebida.tipo_id,

        OfertaRecebida.datahora

    ).all()

    total = sum(
        float(x.valor)
        for x in lista
    )

    return render_template(

        "ofertas/relatorio_ofertas.html",

        lista=lista,

        comunidades=comunidades,

        tipos=tipos,

        total=total,

        data_inicial=data_inicial,

        data_final=data_final,

        comunidade_id=comunidade_id,

        tipo_id=tipo_id

    )

@ofertas_bp.route("/exportar_excel_ofertas")
@login_ofertas_required
def exportar_excel_ofertas():

    grupos, total_geral = obter_dados_relatorio(

        request.args.get("data_inicial"),
        request.args.get("data_final"),
        request.args.get("comunidade"),
        request.args.get("tipo")

    )

    wb = Workbook()

    ws = wb.active

    ws.title = "Relatório"

    linha = 1

    for comunidade, tipos in grupos.items():

        ws.cell(
            linha,
            1,
            comunidade
        )

        linha += 1

        total_comunidade = 0

        for tipo, itens in tipos.items():

            ws.cell(
                linha,
                2,
                tipo
            )

            linha += 1

            ws.append([
                "Data",
                "Pagador",
                "TXID",
                "Valor"
            ])

            linha += 1

            subtotal = 0

            for o in itens:

                ws.append([
                    o.datahora.strftime("%d/%m/%Y %H:%M:%S"),
                    o.pagador or "",
                    o.endtoendid,
                    float(o.valor)
                ])

                subtotal += float(o.valor)

                linha += 1

            ws.append([
                "",
                "",
                f"Subtotal {tipo}",
                subtotal
            ])

            linha += 1

            total_comunidade += subtotal

        ws.append([
            "",
            "",
            "Total Comunidade",
            total_comunidade
        ])

        linha += 2

    ws.append([
        "",
        "",
        "TOTAL GERAL",
        total_geral
    ])

    arquivo = BytesIO()

    wb.save(arquivo)

    arquivo.seek(0)

    return send_file(
        arquivo,
        as_attachment=True,
        download_name="Ofertas.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

from reportlab.platypus import (
    SimpleDocTemplate,
    Paragraph,
    Spacer,
    Table,
    TableStyle
)

from reportlab.lib.styles import getSampleStyleSheet

@ofertas_bp.route("/exportar_pdf_ofertas")
@login_ofertas_required
def exportar_pdf_ofertas():

    grupos, total_geral = obter_dados_relatorio(

        request.args.get("data_inicial"),
        request.args.get("data_final"),
        request.args.get("comunidade"),
        request.args.get("tipo")

    )

    buffer = BytesIO()

    doc = SimpleDocTemplate(buffer)

    styles = getSampleStyleSheet()

    elementos = []

    elementos.append(
        Paragraph(
            "RELATÓRIO DE OFERTAS",
            styles["Title"]
        )
    )

    elementos.append(
        Spacer(1,12)
    )

    for comunidade, tipos in grupos.items():

        elementos.append(
            Paragraph(
                comunidade,
                styles["Heading2"]
            )
        )

        total_comunidade = 0

        for tipo, itens in tipos.items():

            elementos.append(
                Paragraph(
                    tipo,
                    styles["Heading3"]
                )
            )

            dados = [[
                "Data/Hora",
                "Pagador",
                "EndToEndId",
                "Valor"
            ]]

            subtotal = 0

            for o in itens:

                valor = float(o.valor)

                subtotal += valor

                dados.append([
                    o.datahora.strftime("%d/%m/%Y %H:%M:%S"),
                    o.pagador or "",
                    o.endtoendid,
                    f"R$ {valor:.2f}"
                ])

            dados.append([
                "",
                "",
                f"Subtotal {tipo}",
                f"R$ {subtotal:.2f}"
            ])

            tabela = Table(
                dados,
                colWidths=[
                    100,
                    100,
                    250,
                    70
                ]
            )


            tabela.setStyle(
                TableStyle([
                    ("GRID",(0,0),(-1,-1),1,colors.black),
                    ("BACKGROUND",(0,0),(-1,0),colors.lightgrey)
                ])
            )

            elementos.append(tabela)

            elementos.append(
                Spacer(1,8)
            )

            total_comunidade += subtotal

        elementos.append(
            Paragraph(
                f"Total da Comunidade: R$ {total_comunidade:.2f}",
                styles["Heading4"]
            )
        )

        elementos.append(
            Spacer(1,12)
        )

    elementos.append(
        Paragraph(
            f"TOTAL GERAL: R$ {total_geral:.2f}",
            styles["Title"]
        )
    )

    doc.build(elementos)

    buffer.seek(0)

    return send_file(
        buffer,
        as_attachment=False,
        download_name="Ofertas.pdf",
        mimetype="application/pdf"
    )

from collections import OrderedDict

def obter_dados_relatorio(
    data_inicial=None,
    data_final=None,
    comunidade_id=None,
    tipo_id=None
):

    consulta = OfertaRecebida.query

    if not session.get(
        "administrador"
    ):

        consulta = consulta.filter(

            OfertaRecebida.comunidade_id==

            session["comunidade_id"]

        )

    if data_inicial:
        consulta = consulta.filter(
            OfertaRecebida.datahora >=
            datetime.strptime(
                data_inicial,
                "%Y-%m-%d"
            )
        )

    if data_final:
        consulta = consulta.filter(
            OfertaRecebida.datahora <=
            datetime.strptime(
                data_final + " 23:59:59",
                "%Y-%m-%d %H:%M:%S"
            )
        )

    if comunidade_id:
        consulta = consulta.filter(
            OfertaRecebida.comunidade_id ==
            int(comunidade_id)
        )

    if tipo_id:
        consulta = consulta.filter(
            OfertaRecebida.tipo_id ==
            int(tipo_id)
        )

    lista = consulta.order_by(
        OfertaRecebida.comunidade_id,
        OfertaRecebida.tipo_id,
        OfertaRecebida.datahora
    ).all()

    grupos = OrderedDict()

    total_geral = 0

    for o in lista:

        comunidade = (
            o.comunidade.nome
            if o.comunidade else
            "SEM COMUNIDADE"
        )

        tipo = (
            o.tipo.descricao
            if o.tipo else
            "SEM TIPO"
        )

        grupos.setdefault(
            comunidade,
            OrderedDict()
        )

        grupos[comunidade].setdefault(
            tipo,
            []
        )

        grupos[comunidade][tipo].append(o)

        total_geral += float(o.valor)

    return grupos, total_geral

@ofertas_bp.route(
    "/relatorio_analitico",
    methods=["GET", "POST"]
)
@login_ofertas_required

def relatorio_analitico():

    comunidades = Comunidade.query.order_by(
        Comunidade.nome
    ).all()

    tipos = TipoArrecadacao.query.filter_by(
        ativo=True
    ).order_by(
        TipoArrecadacao.descricao
    ).all()

    consulta = OfertaRecebida.query

    if not session.get(
        "administrador"
    ):

        consulta = consulta.filter(

            OfertaRecebida.comunidade_id==

            session["comunidade_id"]

        )

    lista = []

    data_inicial = ""
    data_final = ""
    comunidade_id = ""
    tipo_id = ""

    if request.method == "POST":

        data_inicial = request.form.get(
            "data_inicial"
        )

        data_final = request.form.get(
            "data_final"
        )

        comunidade_id = request.form.get(
            "comunidade_id"
        )

        tipo_id = request.form.get(
            "tipo_id"
        )

        if data_inicial:
            consulta = consulta.filter(
                OfertaRecebida.datahora >= data_inicial
            )

        if data_final:
            consulta = consulta.filter(
                OfertaRecebida.datahora <=
                data_final + " 23:59:59"
            )

        if comunidade_id:
            consulta = consulta.filter(
                OfertaRecebida.comunidade_id ==
                int(comunidade_id)
            )

        if tipo_id:
            consulta = consulta.filter(
                OfertaRecebida.tipo_id ==
                int(tipo_id)
            )

        lista = consulta.order_by(
            OfertaRecebida.datahora.asc()
        ).all()

    total = sum(
        float(x.valor)
        for x in lista
    )

    return render_template(

        "ofertas/relatorio_analitico.html",

        lista=lista,

        total=total,

        comunidades=comunidades,

        tipos=tipos,

        data_inicial=data_inicial,

        data_final=data_final,

        comunidade_id=comunidade_id,

        tipo_id=tipo_id

    )


from flask import send_file, request
from reportlab.platypus import (
    SimpleDocTemplate,
    Table,
    TableStyle
)

from reportlab.lib.pagesizes import landscape, A4


@ofertas_bp.route("/relatorio_analitico_pdf")
@login_ofertas_required
def relatorio_analitico_pdf():

    consulta = OfertaRecebida.query

    if not session.get(
        "administrador"
    ):

        consulta = consulta.filter(

            OfertaRecebida.comunidade_id==

            session["comunidade_id"]

        )

    data_inicial = request.args.get("data_inicial")
    data_final = request.args.get("data_final")
    comunidade_id = request.args.get("comunidade_id")
    tipo_id = request.args.get("tipo_id")

    if data_inicial:
        consulta = consulta.filter(
            OfertaRecebida.datahora >= data_inicial
        )

    if data_final:
        consulta = consulta.filter(
            OfertaRecebida.datahora <= data_final + " 23:59:59"
        )

    if comunidade_id:
        consulta = consulta.filter(
            OfertaRecebida.comunidade_id == int(comunidade_id)
        )

    if tipo_id:
        consulta = consulta.filter(
            OfertaRecebida.tipo_id == int(tipo_id)
        )

    lista = consulta.order_by(
        OfertaRecebida.datahora.asc()
    ).all()

    total = sum(float(o.valor) for o in lista)

    dados = [[
        "Data/Hora",
        "Comunidade",
        "Tipo",
        "Valor"
    ]]

    for o in lista:

        dados.append([

            o.datahora.strftime("%d/%m/%Y %H:%M:%S"),

            o.comunidade.nome if o.comunidade else "",

            o.tipo.descricao if o.tipo else "",

            f"R$ {float(o.valor):.2f}"

        ])

    dados.append([
        "",
        "",
       
        "TOTAL",
        f"R$ {total:.2f}"
    ])

    buffer = BytesIO()

    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4)
    )

    tabela = Table(dados)

    tabela.setStyle(TableStyle([

        ("BACKGROUND", (0, 0), (-1, 0), colors.darkblue),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),

        ("BACKGROUND", (0, -1), (-1, -1), colors.lightgreen),

        ("GRID", (0, 0), (-1, -1), 1, colors.black),

        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),

        ("ALIGN", (4, 1), (4, -1), "RIGHT")

    ]))

    doc.build([tabela])

    buffer.seek(0)

    return send_file(

        buffer,

        download_name="relatorio_analitico_ofertas.pdf",

        as_attachment=True,

        mimetype="application/pdf"

    )

from flask import (
    render_template,
    request,
    redirect,
    url_for,
    flash,
    session
)


@ofertas_bp.route(
    "/login",
    methods=["GET", "POST"]
)
def login():

    comunidades = (
        Comunidade.query
        .filter_by(ativa=True)
        .order_by(Comunidade.nome)
        .all()
    )

    if request.method == "POST":

        comunidade = Comunidade.query.get(
            request.form["comunidade_id"]
        )

        senha = request.form["senha"]

        if not comunidade:
            flash(
                "Comunidade não encontrada.",
                "danger"
            )

        elif comunidade.senha_acesso != senha:
            flash(
                "Senha inválida.",
                "danger"
            )

        else:

            session["ofertas_logado"] = True

            session["comunidade_id"] = comunidade.id

            session["administrador"] = comunidade.administrador

            session["nome_comunidade"] = comunidade.nome

            return redirect(
                url_for(
                    "ofertas.inicio"
                )
            )

    return render_template(
        "ofertas/login.html",
        comunidades=comunidades
    )


@ofertas_bp.route("/logout")
def logout():

    session.pop(
        "ofertas_logado",
        None
    )

    session.pop(
        "comunidade_id",
        None
    )

    session.pop(
        "administrador",
        None
    )

    return redirect(
        url_for(
            "ofertas.login"
        )
    )